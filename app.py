"""Mock Data Factory — Flask application.

Generates realistic mock data based on a user-supplied schema and exports it
as CSV, JSON, XML, SQL, or Excel.

Supports inter-field templates (e.g. `{{first_name|lower}}.{{last_name|lower}}@example.com`)
and schema inference from SQL DDL, JSON samples, or TypeScript interfaces.
"""
from __future__ import annotations

import csv
import io
import json
import logging
import os
import random
import re
import xml.dom.minidom
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta
from typing import Any, Callable

import openpyxl
from faker import Faker
from flask import Flask, jsonify, render_template, request, send_file
from openpyxl.utils import get_column_letter

app = Flask(__name__)
fake = Faker()
logger = logging.getLogger(__name__)

# Hard cap to protect against memory exhaustion / DoS.
MAX_ROWS = int(os.getenv("MOCK_DATA_MAX_ROWS", "100000"))
MAX_PREVIEW_ROWS = 100
MAX_INFER_SOURCE = 32_000  # ~32KB of source text for inference
MAX_FIELDS = 200            # cap field count per schema
MAX_CONTENT_LENGTH = 256 * 1024  # 256KB request body cap

# Reject oversized request bodies before Flask hands them to a route.
app.config["MAX_CONTENT_LENGTH"] = MAX_CONTENT_LENGTH

DEFAULT_SCHEMA = [
    {"name": "id", "type": "Row Number"},
    {"name": "first_name", "type": "First Name"},
    {"name": "last_name", "type": "Last Name"},
    {"name": "email", "type": "Email Address"},
    {"name": "gender", "type": "Gender"},
    {"name": "ip_address", "type": "IP Address v4"},
]

SUPPORTED_FORMATS = {"CSV", "JSON", "XML", "SQL", "EXCEL"}


class SchemaError(ValueError):
    """Raised for any user-facing schema validation failure."""


# ---------- Generators ---------------------------------------------------

def _date_within_last_five_years(_config: dict | None = None) -> str:
    start = datetime.now() - timedelta(days=365 * 5)
    end = datetime.now()
    return fake.date_between(start_date=start, end_date=end).strftime("%Y-%m-%d")


def _custom_list(config: dict | None) -> str:
    if config and config.get("values"):
        return random.choice(config["values"])
    return ""


# Dispatch table — extending the generator is just one new entry.
GENERATORS: dict[str, Callable[[dict | None], Any]] = {
    "Row Number": lambda _c: None,  # Filled by index in generate_data.
    "First Name": lambda _c: fake.first_name(),
    "Last Name": lambda _c: fake.last_name(),
    "Full Name": lambda _c: fake.name(),
    "Email Address": lambda _c: fake.email(),
    "Gender": lambda _c: random.choice(("Male", "Female", "Other")),
    "IP Address v4": lambda _c: fake.ipv4(),
    "Phone Number": lambda _c: fake.phone_number(),
    "City": lambda _c: fake.city(),
    "Country": lambda _c: fake.country(),
    "Date": _date_within_last_five_years,
    "Number": lambda _c: random.randint(1, 1000),
    "Decimal": lambda _c: round(random.uniform(0, 1000), 2),
    "Custom List": _custom_list,
    "Blank/Null": lambda c: None if random.random() < ((c or {}).get("blank_percentage", 0) / 100) else fake.word(),
    # Template values are filled in a second pass after all other fields exist.
    "Template": lambda _c: None,
}


def generate_value(field_type: str, field_config: dict | None = None) -> Any:
    """Generate a single value for a field type."""
    gen = GENERATORS.get(field_type)
    return gen(field_config) if gen else ""


# ---------- Template engine ---------------------------------------------

_TEMPLATE_TOKEN = re.compile(r"\{\{\s*([^}|\s]+)\s*(?:\|\s*([^}]+?)\s*)?\}\}")


def _filter_slug(value: str) -> str:
    return re.sub(r"\W+", "-", str(value)).strip("-").lower()


def _filter_initial(value: str) -> str:
    s = str(value).strip()
    return s[0].upper() if s else ""


def _filter_digits(value: str) -> str:
    return re.sub(r"\D", "", str(value))


TEMPLATE_FILTERS: dict[str, Callable[[Any], Any]] = {
    "lower": lambda v: str(v).lower(),
    "upper": lambda v: str(v).upper(),
    "title": lambda v: str(v).title(),
    "slug": _filter_slug,
    "nospace": lambda v: str(v).replace(" ", ""),
    "initial": _filter_initial,
    "digits": _filter_digits,
    "trim": lambda v: str(v).strip(),
}


def render_template_value(template: str, row: dict[str, Any]) -> str:
    """Replace `{{field|filter}}` tokens in `template` using values from `row`.

    Unknown fields render as empty strings; unknown filters are ignored.
    """
    if not template:
        return ""

    def replace(match: re.Match) -> str:
        name = match.group(1)
        filter_chain = match.group(2)
        value = row.get(name, "")
        if value is None:
            value = ""
        if filter_chain:
            for fname in (f.strip() for f in filter_chain.split("|")):
                fn = TEMPLATE_FILTERS.get(fname)
                if fn:
                    value = fn(value)
        return str(value)

    return _TEMPLATE_TOKEN.sub(replace, template)


# ---------- Core generation ---------------------------------------------

def generate_data(schema: dict, max_rows: int | None = None) -> list[dict]:
    """Generate rows for the schema. `max_rows` (if provided) caps row count."""
    requested = int(schema["num_rows"])
    if requested < 1:
        raise SchemaError("num_rows must be at least 1")

    num_rows = min(requested, max_rows) if max_rows else requested
    num_rows = min(num_rows, MAX_ROWS)

    fields = schema["fields"]
    data: list[dict] = []

    for i in range(num_rows):
        row: dict[str, Any] = {}
        # Pass 1 — non-template fields.
        for field in fields:
            if field["type"] == "Template":
                row[field["name"]] = None
                continue
            value = generate_value(field["type"], field)
            if field["type"] == "Row Number":
                value = i + 1
            blank_pct = field.get("blank_percentage")
            if blank_pct and random.random() < (blank_pct / 100):
                value = None
            row[field["name"]] = value

        # Pass 2 — templates can reference any other field.
        for field in fields:
            if field["type"] != "Template":
                continue
            template = field.get("template", "") or ""
            value = render_template_value(template, row)
            blank_pct = field.get("blank_percentage")
            if blank_pct and random.random() < (blank_pct / 100):
                value = None
            row[field["name"]] = value

        data.append(row)
    return data


# ---------- Formatters --------------------------------------------------

def format_csv(data: list[dict]) -> str:
    output = io.StringIO()
    if data:
        writer = csv.DictWriter(output, fieldnames=list(data[0].keys()))
        writer.writeheader()
        writer.writerows(data)
    return output.getvalue()


def format_json(data: list[dict]) -> str:
    return json.dumps(data, indent=2)


_XML_NAME_INVALID = re.compile(r"[^A-Za-z0-9_\-.]")
_XML_NAME_BAD_START = re.compile(r"^[^A-Za-z_]")


def _sanitize_xml_name(name: str) -> str:
    """Coerce a column name into a syntactically valid XML element name."""
    if not name:
        return "_field"
    safe = _XML_NAME_INVALID.sub("_", name)
    if _XML_NAME_BAD_START.match(safe):
        safe = "_" + safe
    return safe


def format_xml(data: list[dict]) -> str:
    root = ET.Element("records")
    for item in data:
        record = ET.SubElement(root, "record")
        for key, value in item.items():
            field = ET.SubElement(record, _sanitize_xml_name(key))
            field.text = "" if value is None else str(value)

    xml_bytes = ET.tostring(root, encoding="utf-8")
    pretty = xml.dom.minidom.parseString(xml_bytes).toprettyxml(indent="  ")
    if pretty.startswith("<?xml"):
        pretty = "\n".join(pretty.splitlines()[1:])
    return pretty


def _infer_sql_type(column: str, data: list[dict]) -> str:
    for row in data:
        value = row.get(column)
        if value is None:
            continue
        if isinstance(value, bool):
            return "INTEGER"
        if isinstance(value, int):
            return "INTEGER"
        if isinstance(value, float):
            return "REAL"
        if isinstance(value, str) and value.strip():
            try:
                datetime.strptime(value, "%Y-%m-%d")
                return "DATE"
            except ValueError:
                return "TEXT"
        return "TEXT"
    return "TEXT"


def format_sql(data: list[dict], table_name: str = "mock_data") -> str:
    if not data:
        return ""

    columns = list(data[0].keys())
    lines = [
        "-- SQL Data Export",
        f"-- Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "",
        f"CREATE TABLE IF NOT EXISTS {table_name} (",
    ]
    column_defs = [f"    {col} {_infer_sql_type(col, data)}" for col in columns]
    lines.append(",\n".join(column_defs))
    lines.append(");")
    lines.append("")

    column_list = ", ".join(columns)
    for row in data:
        values = []
        for col in columns:
            value = row[col]
            if value is None:
                values.append("NULL")
            elif isinstance(value, bool):
                values.append("1" if value else "0")
            elif isinstance(value, (int, float)):
                values.append(str(value))
            else:
                escaped = str(value).replace("'", "''")
                values.append(f"'{escaped}'")
        lines.append(f"INSERT INTO {table_name} ({column_list}) VALUES ({', '.join(values)});")

    return "\n".join(lines)


def format_excel(data: list[dict]) -> bytes | None:
    if not data:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mock Data"

    headers = list(data[0].keys())
    bold = openpyxl.styles.Font(bold=True)
    max_widths = [len(str(h)) for h in headers]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, header in enumerate(headers, 1):
            value = row_data[header]
            ws.cell(row=row_idx, column=col_idx, value=value)
            if value is not None:
                width = len(str(value))
                if width > max_widths[col_idx - 1]:
                    max_widths[col_idx - 1] = width

    for col_idx, width in enumerate(max_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width + 2

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


FORMATTERS: dict[str, dict] = {
    "CSV": {"format": format_csv, "mime": "text/csv", "ext": "csv"},
    "JSON": {"format": format_json, "mime": "application/json", "ext": "json"},
    "XML": {"format": format_xml, "mime": "application/xml", "ext": "xml"},
    "SQL": {"format": format_sql, "mime": "text/plain", "ext": "sql"},
    "EXCEL": {
        "format": format_excel,
        "mime": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "ext": "xlsx",
    },
}


# ---------- Schema validation -------------------------------------------

def _validate_schema(schema: Any) -> dict:
    if not isinstance(schema, dict):
        raise SchemaError("Schema must be a JSON object")
    for key in ("fields", "num_rows", "format"):
        if key not in schema:
            raise SchemaError(f"Schema missing required key: {key}")
    if not isinstance(schema["fields"], list) or not schema["fields"]:
        raise SchemaError("Schema must contain at least one field")
    if len(schema["fields"]) > MAX_FIELDS:
        raise SchemaError(f"Schema may contain at most {MAX_FIELDS} fields")

    # Per-field shape check — keeps malformed input from reaching the
    # generator (and turning into a 500).
    for idx, field in enumerate(schema["fields"], 1):
        if not isinstance(field, dict):
            raise SchemaError(f"Field #{idx} must be an object")
        name = field.get("name")
        if not isinstance(name, str) or not name.strip():
            raise SchemaError(f"Field #{idx} requires a non-empty name")
        ftype = field.get("type")
        if not isinstance(ftype, str) or not ftype:
            raise SchemaError(f'Field "{name}" requires a "type" string')
        if ftype == "Custom List":
            values = field.get("values")
            if not isinstance(values, list) or not values:
                raise SchemaError(f'Custom List field "{name}" needs a non-empty "values" list')
        if ftype == "Template":
            template = field.get("template")
            if not isinstance(template, str) or not template.strip():
                raise SchemaError(f'Template field "{name}" needs a "template" string')
        pct = field.get("blank_percentage")
        if pct is not None:
            try:
                pct_int = int(pct)
            except (TypeError, ValueError) as exc:
                raise SchemaError(f'blank_percentage on "{name}" must be an integer') from exc
            if not 0 <= pct_int <= 100:
                raise SchemaError(f'blank_percentage on "{name}" must be between 0 and 100')

    fmt = str(schema["format"]).upper()
    if fmt not in SUPPORTED_FORMATS:
        raise SchemaError(f"Unsupported format: {schema['format']}")
    try:
        rows = int(schema["num_rows"])
    except (TypeError, ValueError) as exc:
        raise SchemaError("num_rows must be an integer") from exc
    if rows < 1:
        raise SchemaError("num_rows must be at least 1")
    if rows > MAX_ROWS:
        raise SchemaError(f"num_rows must not exceed {MAX_ROWS}")
    return schema


# ---------- Schema inference --------------------------------------------

# Order matters — first match wins. Use word-boundary or anchored patterns
# to avoid e.g. "address" matching "ip_address" before the IP rule kicks in.
NAME_HEURISTICS: list[tuple[re.Pattern, str]] = [
    (re.compile(r"(?i)^(id|.*_id|.*Id)$"), "Row Number"),
    (re.compile(r"(?i)e?_?mail"), "Email Address"),
    (re.compile(r"(?i)ip[\-_]?address|^ip$|ipv4"), "IP Address v4"),
    (re.compile(r"(?i)phone|tel(ephone)?|mobile|cell"), "Phone Number"),
    (re.compile(r"(?i)(first|given|fore).?name|fname"), "First Name"),
    (re.compile(r"(?i)(last|sur|family).?name|lname"), "Last Name"),
    (re.compile(r"(?i)full[_\-]?name|^name$|display[_\-]?name|^user[_\-]?name$"), "Full Name"),
    (re.compile(r"(?i)gender|sex"), "Gender"),
    (re.compile(r"(?i)city|town"), "City"),
    (re.compile(r"(?i)country|nation"), "Country"),
    (re.compile(r"(?i)created|updated|deleted|modified|date|time|birth|dob|timestamp"), "Date"),
    (re.compile(r"(?i)price|amount|cost|salary|balance|total|decimal|rate"), "Decimal"),
    (re.compile(r"(?i)count|qty|quantity|number|age|rank|score|year"), "Number"),
]


def _name_to_type(name: str, fallback: str = "Full Name") -> str:
    for pattern, dtype in NAME_HEURISTICS:
        if pattern.search(name):
            return dtype
    return fallback


def _sql_type_to_default(sql_type: str) -> str:
    """Coarse fallback when name heuristics don't match — keyed off the SQL type."""
    t = sql_type.upper()
    if any(x in t for x in ("INT", "SERIAL", "BIGINT", "SMALLINT")):
        return "Number"
    if any(x in t for x in ("DECIMAL", "NUMERIC", "FLOAT", "DOUBLE", "REAL", "MONEY")):
        return "Decimal"
    if any(x in t for x in ("DATE", "TIME", "TIMESTAMP")):
        return "Date"
    if any(x in t for x in ("BOOL", "BIT")):
        return "Custom List"
    return "Full Name"


def _ts_type_to_default(ts_type: str) -> str:
    t = ts_type.strip().rstrip(";").strip()
    if t == "number":
        return "Number"
    if t == "boolean":
        return "Custom List"
    if t == "Date":
        return "Date"
    return "Full Name"


_CREATE_TABLE_RE = re.compile(
    r"create\s+table\s+(?:if\s+not\s+exists\s+)?[`\"\[]?(\w+)[`\"\]]?\s*\(",
    re.IGNORECASE,
)
_COLUMN_LINE_RE = re.compile(
    r"^\s*[`\"\[]?(\w+)[`\"\]]?\s+([A-Za-z][A-Za-z0-9_]*(?:\s*\(\s*\d+\s*(?:,\s*\d+\s*)?\))?)",
)


def _extract_paren_body(source: str, open_paren_idx: int) -> str:
    """Return the substring between matching parens starting at `open_paren_idx`."""
    depth = 0
    start = open_paren_idx + 1
    for i in range(open_paren_idx, len(source)):
        ch = source[i]
        if ch == "(":
            depth += 1
        elif ch == ")":
            depth -= 1
            if depth == 0:
                return source[start:i]
    raise SchemaError("Unclosed parenthesis in CREATE TABLE")


def infer_from_sql(source: str) -> list[dict]:
    """Parse a CREATE TABLE statement and return a list of inferred fields."""
    match = _CREATE_TABLE_RE.search(source)
    if not match:
        raise SchemaError("Could not find a CREATE TABLE statement")
    # match.end() - 1 points to the opening "(" of the column list.
    body = _extract_paren_body(source, match.end() - 1)

    fields: list[dict] = []
    depth = 0
    buf: list[str] = []
    raw_lines: list[str] = []
    for ch in body:
        if ch == "(":
            depth += 1
        elif ch == ")":
            depth -= 1
        if ch == "," and depth == 0:
            raw_lines.append("".join(buf))
            buf = []
        else:
            buf.append(ch)
    if buf:
        raw_lines.append("".join(buf))

    skip_keywords = {"primary", "foreign", "unique", "constraint", "key", "check", "index"}

    for raw in raw_lines:
        line = raw.strip()
        if not line:
            continue
        first_word = line.split()[0].strip("`\"[]").lower() if line.split() else ""
        if first_word in skip_keywords:
            continue
        m = _COLUMN_LINE_RE.match(line)
        if not m:
            continue
        col_name = m.group(1)
        sql_type = m.group(2)
        fallback = _sql_type_to_default(sql_type)
        dtype = _name_to_type(col_name, fallback)
        field: dict[str, Any] = {"name": col_name, "type": dtype}
        if dtype == "Custom List" and fallback == "Custom List":
            field["values"] = ["true", "false"]
        fields.append(field)

    if not fields:
        raise SchemaError("CREATE TABLE statement had no recognizable columns")
    return fields


_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
_IPV4_RE = re.compile(r"^(\d{1,3}\.){3}\d{1,3}$")
_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}")


def _json_value_to_type(name: str, value: Any) -> dict:
    if isinstance(value, bool):
        return {"name": name, "type": "Custom List", "values": ["true", "false"]}
    if isinstance(value, int):
        return {"name": name, "type": _name_to_type(name, "Number")}
    if isinstance(value, float):
        return {"name": name, "type": _name_to_type(name, "Decimal")}
    if value is None:
        return {"name": name, "type": _name_to_type(name)}
    if isinstance(value, str):
        if _EMAIL_RE.match(value):
            return {"name": name, "type": "Email Address"}
        if _IPV4_RE.match(value):
            return {"name": name, "type": "IP Address v4"}
        if _DATE_RE.match(value):
            return {"name": name, "type": "Date"}
        return {"name": name, "type": _name_to_type(name)}
    if isinstance(value, list):
        return {"name": name, "type": _name_to_type(name)}
    if isinstance(value, dict):
        return {"name": name, "type": _name_to_type(name)}
    return {"name": name, "type": "Full Name"}


def infer_from_json(source: str) -> list[dict]:
    """Parse a JSON object or array of objects and return inferred fields."""
    try:
        parsed = json.loads(source)
    except json.JSONDecodeError as exc:
        raise SchemaError(f"Invalid JSON: {exc.msg}") from exc

    if isinstance(parsed, list):
        if not parsed:
            raise SchemaError("JSON array is empty")
        sample = parsed[0]
    else:
        sample = parsed

    if not isinstance(sample, dict):
        raise SchemaError("JSON must be an object or an array of objects")

    return [_json_value_to_type(k, v) for k, v in sample.items()]


_TS_INTERFACE_RE = re.compile(
    r"(?:interface|type)\s+\w+\s*=?\s*\{([^}]+)\}",
    re.DOTALL,
)
_TS_FIELD_RE = re.compile(r"^\s*(\w+)\s*\??\s*:\s*(.+?)\s*$")


def infer_from_typescript(source: str) -> list[dict]:
    """Parse a single TypeScript interface or type alias and return inferred fields."""
    match = _TS_INTERFACE_RE.search(source)
    if not match:
        raise SchemaError("Could not find a TypeScript interface or type")

    # Split on semicolons OR newlines so single-line and multi-line bodies work.
    body = match.group(1)
    parts = re.split(r"[;\n]", body)

    fields: list[dict] = []
    for part in parts:
        line = part.strip()
        if not line or line.startswith("//"):
            continue
        m = _TS_FIELD_RE.match(line)
        if not m:
            continue
        name, ts_type = m.group(1), m.group(2)
        fallback = _ts_type_to_default(ts_type)
        dtype = _name_to_type(name, fallback)
        field: dict[str, Any] = {"name": name, "type": dtype}
        if dtype == "Custom List" and fallback == "Custom List":
            field["values"] = ["true", "false"]
        fields.append(field)

    if not fields:
        raise SchemaError("Interface body had no recognizable fields")
    return fields


INFER_HANDLERS: dict[str, Callable[[str], list[dict]]] = {
    "sql": infer_from_sql,
    "json": infer_from_json,
    "typescript": infer_from_typescript,
    "ts": infer_from_typescript,
}


# ---------- Routes ------------------------------------------------------

@app.errorhandler(413)
def request_too_large(_e):
    return jsonify({"error": f"Request body exceeds {MAX_CONTENT_LENGTH} bytes"}), 413


@app.route("/")
def index():
    return render_template("index.html", default_schema=DEFAULT_SCHEMA)


@app.route("/preview", methods=["POST"])
def preview():
    try:
        schema = _validate_schema(request.get_json(silent=True))
    except SchemaError as exc:
        return jsonify({"error": str(exc)}), 400

    try:
        data = generate_data(schema, MAX_PREVIEW_ROWS)
        fmt = schema["format"].upper()
        if fmt in ("CSV", "EXCEL"):
            field_names = [field["name"] for field in schema["fields"]]
            return jsonify({"data": data, "field_order": field_names}), 200

        formatter = FORMATTERS[fmt]["format"]
        return formatter(data), 200, {"Content-Type": "text/plain"}
    except SchemaError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception:
        logger.exception("Unexpected error generating preview")
        return jsonify({"error": "Internal server error"}), 500


@app.route("/generate", methods=["POST"])
def generate():
    try:
        schema = _validate_schema(request.get_json(silent=True))
    except SchemaError as exc:
        return jsonify({"error": str(exc)}), 400

    try:
        if request.args.get("preview") == "true":
            data = generate_data(schema)
            field_names = [field["name"] for field in schema["fields"]]
            return jsonify({"data": data, "field_order": field_names})

        data = generate_data(schema)
        fmt = schema["format"].upper()
        info = FORMATTERS[fmt]
        output = info["format"](data)
        if isinstance(output, str):
            output = output.encode("utf-8")

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return send_file(
            io.BytesIO(output or b""),
            mimetype=info["mime"],
            as_attachment=True,
            download_name=f"mock_data_{timestamp}.{info['ext']}",
        )
    except SchemaError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception:
        logger.exception("Unexpected error generating data")
        return jsonify({"error": "Internal server error"}), 500


@app.route("/infer-schema", methods=["POST"])
def infer_schema():
    """Infer a schema from SQL DDL, a JSON sample, or a TypeScript interface."""
    payload = request.get_json(silent=True) or {}
    kind = str(payload.get("kind", "")).lower().strip()
    source = payload.get("source", "")
    if not isinstance(source, str) or not source.strip():
        return jsonify({"error": "Provide source text to infer from"}), 400
    if len(source) > MAX_INFER_SOURCE:
        return jsonify({"error": f"Source exceeds {MAX_INFER_SOURCE} characters"}), 400

    handler = INFER_HANDLERS.get(kind)
    if not handler:
        return jsonify({"error": f"Unknown kind: {kind}. Use sql, json, or typescript."}), 400

    try:
        fields = handler(source)
    except SchemaError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception:
        logger.exception("Unexpected error inferring schema")
        return jsonify({"error": "Internal server error"}), 500

    return jsonify({"fields": fields}), 200


if __name__ == "__main__":
    debug = os.getenv("FLASK_DEBUG", "0") == "1"
    app.run(debug=debug, host="127.0.0.1", port=int(os.getenv("PORT", "5000")))
